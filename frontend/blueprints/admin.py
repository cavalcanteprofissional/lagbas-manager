# Admin blueprint - Administrative functions
import json
import io
import jwt
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response
from openpyxl import Workbook

from utils.supabase_utils import get_admin_client
from blueprints.helpers import get_user_id, is_admin, get_user_role, get_habilitar_abas

admin_bp = Blueprint('admin', __name__)


def validate_admin_token():
    """Valida o token JWT para operações admin"""
    user_id = get_user_id()
    token = session.get("jwt_token")
    
    if not token:
        flash("Sessão inválida. Faça login novamente.", "danger")
        return None, redirect(url_for("auth.login"))
    
    try:
        from flask import current_app
        decoded = jwt.decode(token, current_app.secret_key, algorithms=["HS256"])
        if decoded.get("user_id") != user_id:
            flash("Acesso não autorizado.", "danger")
            return None, redirect(url_for("dashboard"))
    except:
        flash("Token inválido.", "danger")
        return None, redirect(url_for("auth.login"))
    
    return user_id, None


@admin_bp.route("/admin")
def panel():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    client = get_admin_client()
    users_response = client.table("perfil").select("*").execute()
    users = users_response.data or []
    
    for user in users:
        user_id = user.get("id")
        
        cilindro_count = client.table("cilindro").select("id", count="exact").eq("user_id", user_id).execute()
        elemento_count = client.table("elemento").select("id", count="exact").eq("user_id", user_id).execute()
        amostra_count = client.table("amostra").select("id", count="exact").eq("user_id", user_id).execute()
        pressao_count = client.table("pressao").select("id", count="exact").eq("user_id", user_id).execute()
        
        user["nome"] = user.get("nome") or user.get("email") or user_id
        user["cilindros"] = cilindro_count.count or 0
        user["elementos"] = elemento_count.count or 0
        user["amostras"] = amostra_count.count or 0
        user["pressoes"] = pressao_count.count or 0
        if user.get("role") == "admin":
            user["habilitar_abas"] = {"cilindro": True, "pressao": True, "elemento": True, "amostra": True, "historico": True}
        else:
            user["habilitar_abas"] = get_habilitar_abas(user["id"])
    
    users = sorted(users, key=lambda x: (x.get("nome") or x.get("email") or "").lower())
    
    return render_template("admin.html", users=users)


@admin_bp.route("/admin/toggle-user", methods=["POST"])
def toggle_user():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    ativo = request.form.get("ativo") == "true"
    
    if target_user_id == get_user_id():
        flash("Você não pode desativar seu próprio usuário.", "warning")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    client.table("perfil").update({"ativo": ativo}).eq("id", target_user_id).execute()
    flash(f"Usuário {'ativado' if ativo else 'desativado'} com sucesso!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/set-role", methods=["POST"])
def set_role():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    role = request.form.get("role")
    
    if role not in ["admin", "usuario"]:
        flash("Função inválida.", "danger")
        return redirect(url_for("admin.panel"))
    
    if target_user_id == get_user_id():
        flash("Você não pode alterar sua própria função.", "warning")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    client.table("perfil").update({"role": role}).eq("id", target_user_id).execute()
    flash(f"Função do usuário alterada para {role}!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/delete-user", methods=["POST"])
def delete_user():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    
    if target_user_id == get_user_id():
        flash("Você não pode excluir seu próprio usuário.", "warning")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    client.table("cilindro").delete().eq("user_id", target_user_id).execute()
    client.table("elemento").delete().eq("user_id", target_user_id).execute()
    client.table("amostra").delete().eq("user_id", target_user_id).execute()
    client.table("perfil").delete().eq("id", target_user_id).execute()
    
    flash("Usuário e todos os seus dados foram excluídos!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/update-habilitar-abas", methods=["POST"])
def update_habilitar_abas():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    target_user_id = request.form.get("user_id")
    aba = request.form.get("aba")
    habilitar = request.form.get("habilitar") == "true"
    
    if not target_user_id or not aba:
        flash("Parâmetros inválidos.", "danger")
        return redirect(url_for("admin.panel"))
    
    if aba not in ["cilindro", "pressao", "elemento", "amostra", "historico"]:
        flash("Aba inválida.", "danger")
        return redirect(url_for("admin.panel"))
    
    client = get_admin_client()
    
    habilitar_abas = {"cilindro": False, "pressao": False, "elemento": False, "amostra": False, "historico": False}
    if perfil.data and perfil.data[0].get("habilitar_abas"):
        habilitar_abas = perfil.data[0].get("habilitar_abas")
    
    habilitar_abas[aba] = habilitar
    
    client.table("perfil").update({"habilitar_abas": habilitar_abas}).eq("id", target_user_id).execute()
    
    acao = "habilitada" if habilitar else "desabilitada"
    nome_aba = {"cilindro": "Cilindros", "pressao": "Pressão", "elemento": "Elementos", "amostra": "Amostras", "historico": "Histórico"}.get(aba, aba)
    flash(f"Aba {nome_aba} {acao} com sucesso!", "success")
    
    client = get_admin_client()
    client.table("perfil").update({"habilitar_abas": habilitar_abas}).eq("id", target_user_id).execute()
    
    flash("Permissões de acesso atualizadas!", "success")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/admin/user-data/<target_user_id>", methods=["GET"])
def user_data(target_user_id):
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    client = get_admin_client()
    
    page = int(request.args.get("page", 1))
    per_page = 20
    
    cilindro_total = client.table("cilindro").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    elementos_total = client.table("elemento").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    amostras_total = client.table("amostra").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    pressoes_total = client.table("pressao").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    
    historico_offset = (page - 1) * per_page
    historico_log = client.table("historico_log").select(
        "tipo, acao, nome, created_at"
    ).eq("user_id", target_user_id).order("created_at", desc=True).range(historico_offset, historico_offset + per_page - 1).execute().data or []
    
    historico_total = client.table("historico_log").select("*", count="exact").eq("user_id", target_user_id).execute().count or 0
    
    history = [{
        "tipo": h.get("tipo"),
        "acao": h.get("acao"),
        "nome": h.get("nome"),
        "data": h.get("created_at")
    } for h in historico_log]
    
    perfil = client.table("perfil").select("*").eq("id", target_user_id).execute().data
    target_user = perfil[0] if perfil else {"id": target_user_id, "role": "unknown"}
    
    habilitar_abas = get_habilitar_abas(target_user_id) if target_user.get("role") != "admin" else {"cilindro": True, "pressao": True, "elemento": True, "amostra": True, "historico": True}
    
    return render_template(
        "admin_user_data.html",
        target_user=target_user,
        cilindro_total=cilindro_total,
        elementos_total=elementos_total,
        amostras_total=amostras_total,
        pressoes_total=pressoes_total,
        habilitar_abas=habilitar_abas,
        history=history,
        historico_total=historico_total,
        page=page,
        per_page=per_page
    )


@admin_bp.route("/admin/export")
def export_data():
    if not is_admin():
        flash("Acesso restrito a administradores.", "danger")
        return redirect(url_for("dashboard"))
    
    user_id, error = validate_admin_token()
    if error:
        return error
    
    formato = request.args.get("formato", "json").lower()
    
    if formato not in ["csv", "json", "excel", "md"]:
        flash("Formato inválido.", "danger")
        return redirect(url_for("dashboard"))
    
    client = get_admin_client()
    
    cilindro_data = client.table("cilindro").select("*").execute().data or []
    elementos_data = client.table("elemento").select("*").execute().data or []
    amostras_data = client.table("amostra").select("*").execute().data or []
    pressoes_data = client.table("pressao").select("*").execute().data or []
    usuarios_data = client.table("perfil").select("id,email,nome").execute().data or []
    
    usuarios_dict = {u.get("id"): u for u in usuarios_data}
    
    for c in cilindro_data:
        uid = c.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            c["usuario_email"] = u.get("email", "")
            c["usuario_nome"] = u.get("nome", "")
    
    for e in elementos_data:
        uid = e.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            e["usuario_email"] = u.get("email", "")
            e["usuario_nome"] = u.get("nome", "")
    
    cilindro_dict = {c.get("id"): c.get("codigo") for c in cilindro_data}
    elemento_dict = {e.get("id"): e.get("nome") for e in elementos_data}
    
    for a in amostras_data:
        uid = a.get("user_id")
        if uid:
            u = usuarios_dict.get(uid, {})
            a["usuario_email"] = u.get("email", "")
            a["usuario_nome"] = u.get("nome", "")
        a["cilindro_codigo"] = cilindro_dict.get(a.get("cilindro_id"), "")
        a["elemento_nome"] = elemento_dict.get(a.get("elemento_id"), "")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if formato == "json":
        data = {
            "exportado_em": datetime.now().isoformat(),
            "cilindros": cilindro_data,
            "elementos": elementos_data,
            "amostras": amostras_data,
            "temperaturas": pressoes_data
        }
        response = make_response(json.dumps(data, indent=2, default=str))
        response.headers["Content-Disposition"] = f"attachment; filename=labgas_export_{timestamp}.json"
        response.headers["Content-Type"] = "application/json"
        return response
    
    elif formato == "csv":
        output = io.StringIO()
        
        output.write("# CILINDROS\n")
        if cilindro_data:
            headers = ["id", "codigo", "data_compra", "data_inicio_consumo", "data_fim", 
                      "gas_kg", "litros_equivalentes", "custo", "status", 
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in cilindro_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# ELEMENTOS\n")
        if elementos_data:
            headers = ["id", "nome", "consumo_lpm", "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in elementos_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# AMOSTRAS\n")
        if amostras_data:
            headers = ["id", "data", "tempo_chama", "cilindro_id", "cilindro_codigo", 
                      "elemento_id", "elemento_nome", "quantidade_amostras",
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in amostras_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        output.write("\n# TEMPERATURAS\n")
        if pressoes_data:
            for t in pressoes_data:
                uid = t.get("user_id")
                if uid:
                    u = usuarios_dict.get(uid, {})
                    t["usuario_email"] = u.get("email", "")
                    t["usuario_nome"] = u.get("nome", "")
                t["cilindro_codigo"] = cilindro_dict.get(t.get("cilindro_id"), "")
            headers = ["id", "cilindro_id", "cilindro_codigo", "temperatura", "data", "hora",
                      "usuario_email", "usuario_nome", "created_at"]
            output.write(",".join(headers) + "\n")
            for row in pressoes_data:
                values = [str(row.get(h, "")) for h in headers]
                output.write(",".join(values) + "\n")
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=labgas_export_{timestamp}.csv"
        response.headers["Content-Type"] = "text/csv"
        return response
    
    elif formato == "excel":
        wb = Workbook()
        
        ws_cilindros = wb.active
        ws_cilindros.title = "Cilindros"
        if cilindro_data:
            headers = ["ID", "Código", "Data Compra", "Data Início", "Data Fim", 
                      "Gas (kg)", "Litros", "Custo", "Status", 
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_cilindros.append(headers)
            for row in cilindro_data:
                ws_cilindros.append([
                    row.get("id"), row.get("codigo"), row.get("data_compra"),
                    row.get("data_inicio_consumo"), row.get("data_fim"),
                    row.get("gas_kg"), row.get("litros_equivalentes"), row.get("custo"),
                    row.get("status"), 
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_elementos = wb.create_sheet("Elementos")
        if elementos_data:
            headers = ["ID", "Nome", "Consumo (L/min)", "Usuário Email", "Usuário Nome", "Criado em"]
            ws_elementos.append(headers)
            for row in elementos_data:
                ws_elementos.append([
                    row.get("id"), row.get("nome"), row.get("consumo_lpm"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_amostras = wb.create_sheet("Amostras")
        if amostras_data:
            headers = ["ID", "Data", "Tempo Chama", "Cilindro ID", "Cilindro Código",
                      "Elemento ID", "Elemento Nome", "Qtd Amostras",
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_amostras.append(headers)
            for row in amostras_data:
                ws_amostras.append([
                    row.get("id"), row.get("data"), row.get("tempo_chama"),
                    row.get("cilindro_id"), row.get("cilindro_codigo"),
                    row.get("elemento_id"), row.get("elemento_nome"),
                    row.get("quantidade_amostras"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        ws_pressoes = wb.create_sheet("Pressoes")
        if pressoes_data:
            for t in pressoes_data:
                t["cilindro_codigo"] = cilindro_dict.get(t.get("cilindro_id"), "")
            headers = ["ID", "Cilindro ID", "Cilindro Código", "Temperatura (°C)", "Data", "Hora",
                      "Usuário Email", "Usuário Nome", "Criado em"]
            ws_temperaturas.append(headers)
            for row in pressoes_data:
                ws_temperaturas.append([
                    row.get("id"), row.get("cilindro_id"), row.get("cilindro_codigo"),
                    row.get("temperatura"), row.get("data"), row.get("hora"),
                    row.get("usuario_email"), row.get("usuario_nome"), row.get("created_at")
                ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=labgas_export_{timestamp}.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    
    elif formato == "md":
        md_output = io.StringIO()
        
        md_output.write(f"# LabGas Manager - Exportação\n\n")
        md_output.write(f"**Exportado em:** {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
        md_output.write(f"**Total:** {len(cilindro_data)} Cilindros | {len(pressoes_data)} Pressoes | {len(elementos_data)} Elementos | {len(amostras_data)} Amostras\n\n")
        
        md_output.write("## Cilindros\n\n")
        if cilindro_data:
            md_output.write("| ID | Código | Status | Gas (kg) | Custo | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|\n")
            for row in cilindro_data:
                md_output.write(f"| {row.get('id')} | {row.get('codigo')} | {row.get('status')} | {row.get('gas_kg')} | R${row.get('custo')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhum cilindro encontrado.*\n\n")
        
        md_output.write("\n## Pressoes\n\n")
        if pressoes_data:
            for t in pressoes_data:
                t["cilindro_codigo"] = cilindro_dict.get(t.get("cilindro_id"), "")
            md_output.write("| ID | Cilindro | Temperatura | Data | Hora | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|\n")
            for row in pressoes_data:
                md_output.write(f"| {row.get('id')} | {row.get('cilindro_codigo')} | {row.get('temperatura')}°C | {row.get('data')} | {row.get('hora')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhum registro de temperatura encontrado.*\n\n")
        
        md_output.write("\n## Elementos\n\n")
        if elementos_data:
            md_output.write("| ID | Nome | Consumo (L/min) | Usuário |\n")
            md_output.write("|---|---|---|---|\n")
            for row in elementos_data:
                md_output.write(f"| {row.get('id')} | {row.get('nome')} | {row.get('consumo_lpm')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhum elemento encontrado.*\n\n")
        
        md_output.write("\n## Amostras\n\n")
        if amostras_data:
            md_output.write("| ID | Data | Tempo | Cilindro | Elemento | Qtd | Usuário |\n")
            md_output.write("|---|---|---|---|---|---|---|\n")
            for row in amostras_data:
                md_output.write(f"| {row.get('id')} | {row.get('data')} | {row.get('tempo_chama')} | {row.get('cilindro_codigo')} | {row.get('elemento_nome')} | {row.get('quantidade_amostras')} | {row.get('usuario_email')} |\n")
        else:
            md_output.write("*Nenhuma amostra encontrada.*\n\n")
        
        response = make_response(md_output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=labgas_export_{timestamp}.md"
        response.headers["Content-Type"] = "text/markdown"
        return response
    
    flash("Formato não suportado.", "danger")
    return redirect(url_for("dashboard"))
